import pywhatkit as kit
from datetime import datetime, timedelta
from openpyxl import Workbook,load_workbook

def mesaj_gonder(phone_number, message, dakika_sonra=2, wait_time=10):
    """
    Belirli bir süre sonra WhatsApp mesajı gönderir.

    Parametreler:
        phone_number (str): Alıcının telefon numarası (uluslararası formatta).
        message (str): Gönderilecek mesaj içeriği.
        dakika_sonra (int): Mesajın gönderileceği süre (dakika olarak, şu andan itibaren).
        wait_time (int): WhatsApp Web'in yüklenmesi için bekleme süresi (saniye olarak).
    """
    # Şu anki zamanı al ve gönderim saatini belirle
    now = datetime.now()
    send_time = now + timedelta(minutes=dakika_sonra)
    hour = send_time.hour
    minute = send_time.minute

    try:
        # Mesaj gönder
        kit.sendwhatmsg(
            phone_number,
            message,
            hour,
            minute,
            wait_time=wait_time
        )
        print(f"Mesaj {hour}:{minute}'de gönderilecek.")
    except Exception as e:
        print(f"Mesaj gönderilirken bir hata oluştu: {e}")
"""Exceli okyom burda """
wb=load_workbook("tele.xlsx")
ws=wb.active
"""döngü yapacam"""
for s in range(1,ws.max_row+1):
    a=ws.cell(s,1).value
    a=a.replace(" ","")
    print(a)
    
    mesaj_gonder(
    a,
    "Bu mesaj Python ile gönderildi.",
    dakika_sonra=2,
    wait_time=10
 )


